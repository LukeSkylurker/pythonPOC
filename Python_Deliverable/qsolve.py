from openpyxl import Workbook
import openpyxl

wb = openpyxl.load_workbook('rawintake.xlsx')
ws1 = wb.active
wb2 = openpyxl.load_workbook('output.xlsx')
ws2 = wb2.active

num_row = 1

for row in ws1:
    for cell in row:
        if cell.value == "MENS APPAREL" or cell.value == "LADIES APPAREL" or cell.value == "CHILDRENS APPAREL":
            num_row = cell.row
            print(num_row)
            one = ws1.cell(row=cell.row, column=1).value
            two = ws1.cell(row=cell.row, column=2).value
            three = ws1.cell(row=cell.row, column=3).value
            four = ws1.cell(row=cell.row, column=4).value
            five = ws1.cell(row=cell.row, column=5).value
            six = ws1.cell(row=cell.row, column=6).value
            seven = ws1.cell(row=cell.row, column=7).value
            eight = ws1.cell(row=cell.row, column=8).value
            nine = ws1.cell(row=cell.row, column=9).value
            ten = ws1.cell(row=cell.row, column=10).value
            all_num = [one, two, three,four, five, six, seven, eight]
            print(all_num)
            ws2.append(all_num)
wb2.save("output.xlsx")